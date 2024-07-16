import tkinter as tk
from tkinter import messagebox
from PIL import Image, ImageTk
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import datetime
import os

class BarrelCleaningApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Registro de Limpieza de Barriles")
        self.root.state('zoomed')

        self.create_widgets()
        self.setup_excel()
        self.calculate_totals(None)
        self.update_color()

    def create_widgets(self):
        font_large_bold = ("Helvetica", 10, "bold")
        font_title = ("Helvetica", 12, "bold")

        titles_line1 = [
            "Fecha de limpieza", "50 lts", "30 lts", "20 lts", "Totales",
            "Enjuague\nAgua 70°C\n(X2)", "Detergente\nAlcalino\n(Ciclo 3 min.)",
            "Enjuague\nAgua Fria\n(X4)", "Acido\nFosforico", "Acido\nPer-acetico\n(Ciclo 2 min.)"
        ]

        titles_line2 = [
            "Lavado de\nLanzas", "Responsable"
        ]

        for idx, title in enumerate(titles_line1):
            label = tk.Label(self.root, text=title, font=font_title)
            label.grid(row=0, column=idx, padx=5, pady=5, sticky="nsew")

        for idx, title in enumerate(titles_line2):
            tk.Label(self.root, text=title, font=font_title).grid(row=0, column=len(titles_line1) + idx, padx=5, pady=5, sticky="nsew")

        self.fecha_entry = tk.Entry(self.root, width=12)
        self.fecha_entry.grid(row=1, column=0, padx=5, pady=5)
        self.fecha_entry.insert(0, datetime.datetime.now().strftime("%Y-%m-%d"))

        self.barril_50_entry = tk.Entry(self.root, width=5)
        self.barril_50_entry.grid(row=1, column=1, padx=5, pady=5)
        self.barril_50_entry.bind('<FocusOut>', self.calculate_totals)

        self.barril_30_entry = tk.Entry(self.root, width=5)
        self.barril_30_entry.grid(row=1, column=2, padx=5, pady=5)
        self.barril_30_entry.bind('<FocusOut>', self.calculate_totals)

        self.barril_20_entry = tk.Entry(self.root, width=5)
        self.barril_20_entry.grid(row=1, column=3, padx=5, pady=5)
        self.barril_20_entry.bind('<FocusOut>', self.calculate_totals)

        self.total_label = tk.Label(self.root, text="", font=("Helvetica", 10, "bold"))
        self.total_label.grid(row=1, column=4, padx=5, pady=5)

        self.enjuague_70_var = tk.BooleanVar()
        self.enjuague_70_cb = tk.Checkbutton(self.root, variable=self.enjuague_70_var, command=self.update_color,
                                             font=("Helvetica", 10, "bold"), width=2)
        self.enjuague_70_cb.grid(row=1, column=5, padx=5, pady=5)

        self.detergente_var = tk.BooleanVar()
        self.detergente_cb = tk.Checkbutton(self.root, variable=self.detergente_var, command=self.update_color,
                                            font=("Helvetica", 10, "bold"), width=2)
        self.detergente_cb.grid(row=1, column=6, padx=5, pady=5)

        self.enjuague_frio_var = tk.BooleanVar()
        self.enjuague_frio_cb = tk.Checkbutton(self.root, variable=self.enjuague_frio_var, command=self.update_color,
                                               font=("Helvetica", 10, "bold"), width=2)
        self.enjuague_frio_cb.grid(row=1, column=7, padx=5, pady=5)

        self.acido_fosforico_var = tk.BooleanVar()
        self.acido_fosforico_cb = tk.Checkbutton(self.root, variable=self.acido_fosforico_var, command=self.update_color,
                                                 font=("Helvetica", 10, "bold"), width=2)
        self.acido_fosforico_cb.grid(row=1, column=8, padx=5, pady=5)

        self.acido_peracetico_var = tk.BooleanVar()
        self.acido_peracetico_cb = tk.Checkbutton(self.root, variable=self.acido_peracetico_var, command=self.update_color,
                                                  font=("Helvetica", 10, "bold"), width=2)
        self.acido_peracetico_cb.grid(row=1, column=9, padx=5, pady=5)

        self.lanzas_var = tk.BooleanVar()
        self.lanzas_cb = tk.Checkbutton(self.root, variable=self.lanzas_var, command=self.update_color,
                                        font=("Helvetica", 10, "bold"), width=2)
        self.lanzas_cb.grid(row=1, column=10, padx=5, pady=5)

        self.responsable_entry = tk.Entry(self.root, width=15)
        self.responsable_entry.grid(row=1, column=11, padx=5, pady=5)

        self.submit_button = tk.Button(self.root, text="Registrar", command=self.save_data, font=font_large_bold)
        self.submit_button.grid(row=2, column=0, columnspan=len(titles_line1) + len(titles_line2), pady=10)

        self.logo_image = ImageTk.PhotoImage(Image.open("C:/Logo.png"))
        self.logo_label = tk.Label(self.root, image=self.logo_image)
        self.logo_label.grid(row=3, column=0, columnspan=2, sticky="w", pady=10)

        for idx in range(len(titles_line1) + len(titles_line2)):
            self.root.columnconfigure(idx, weight=1)

    def update_color(self):
        self._update_checkbox_color(self.enjuague_70_var, self.enjuague_70_cb)
        self._update_checkbox_color(self.detergente_var, self.detergente_cb)
        self._update_checkbox_color(self.enjuague_frio_var, self.enjuague_frio_cb)
        self._update_checkbox_color(self.acido_fosforico_var, self.acido_fosforico_cb)
        self._update_checkbox_color(self.acido_peracetico_var, self.acido_peracetico_cb)
        self._update_checkbox_color(self.lanzas_var, self.lanzas_cb)

    def _update_checkbox_color(self, var, checkbox):
        if var.get():
            checkbox.config(bg="green", activebackground="green")
        else:
            checkbox.config(bg=self.root.cget("bg"), activebackground=self.root.cget("bg"))

    def setup_excel(self):
        documents_folder = os.path.expanduser('~\\Documents')
        registros_folder = os.path.join(documents_folder, "Registros Fabrica")

        if not os.path.exists(registros_folder):
            try:
                os.makedirs(registros_folder)
                messagebox.showinfo("Carpeta creada", "Se ha creado la carpeta 'Registros Fabrica'.")
            except OSError as e:
                messagebox.showerror("Error", f"No se pudo crear la carpeta 'Registros Fabrica': {e}")
                self.root.destroy()
                return

        self.filename = os.path.join(registros_folder, "registro_limpieza_barriles.xlsx")

        if not os.path.exists(self.filename):
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.append(["Fecha de limpieza", "50 lts", "30 lts", "20 lts", "Totales",
                            "Enjuague 70°C", "Detergente Alcalino", "Enjuague Agua Fria", "Acido Fosforico",
                            "Acido Per-acetico", "Lavado de Lanzas", "Responsable"])

            title_font = Font(bold=True)
            cell_fill = PatternFill(fill_type='solid', fgColor="FFE699")
            alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            for cell in self.ws[1]:
                cell.font = title_font
                cell.fill = cell_fill
                cell.alignment = alignment

            for col in self.ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                self.ws.column_dimensions[column].width = adjusted_width

            self.wb.save(self.filename)
        else:
            self.wb = load_workbook(self.filename)
            self.ws = self.wb.active

    def save_data(self):
        fecha_str = self.fecha_entry.get()
        try:
            fecha = datetime.datetime.strptime(fecha_str, "%Y-%m-%d").date()
        except ValueError:
            messagebox.showerror("Error", "La fecha debe estar en formato YYYY-MM-DD.")
            return

        try:
            barril_50 = int(self.barril_50_entry.get()) if self.barril_50_entry.get() else 0
            barril_30 = int(self.barril_30_entry.get()) if self.barril_30_entry.get() else 0
            barril_20 = int(self.barril_20_entry.get()) if self.barril_20_entry.get() else 0
        except ValueError:
            messagebox.showerror("Error", "Las cantidades de barriles deben ser números enteros.")
            return

        total_barriles = barril_50 + barril_30 + barril_20

        enjuague_70 = "Sí" if self.enjuague_70_var.get() else "No"
        detergente = "Sí" if self.detergente_var.get() else "No"
        enjuague_frio = "Sí" if self.enjuague_frio_var.get() else "No"
        acido_fosforico = "Sí" if self.acido_fosforico_var.get() else "No"
        acido_peracetico = "Sí" if self.acido_peracetico_var.get() else "No"
        lanzas = "Sí" if self.lanzas_var.get() else "No"
        responsable = self.responsable_entry.get()

        print(f"Debug: Registrando fecha: {fecha}, Barril 50 lts: {barril_50}, Barril 30 lts: {barril_30}, Barril 20 lts: {barril_20}, Total: {total_barriles}, Responsable: {responsable}")

        for row in self.ws.iter_rows(min_row=2, values_only=True):
            row_date = row[0].date() if isinstance(row[0], datetime.datetime) else row[0]
            print(f"Debug: Verificando fila: {row}")
            if row_date == fecha:
                messagebox.showerror("Error", "Ya existe un registro con esta fecha.")
                return

        new_data = [fecha, barril_50, barril_30, barril_20, total_barriles,
                    enjuague_70, detergente, enjuague_frio, acido_fosforico,
                    acido_peracetico, lanzas, responsable]

        self.ws.append(new_data)
        self.sort_excel_data()
        self.wb.save(self.filename)

        messagebox.showinfo("Registro guardado", "Se ha guardado el registro de limpieza de barriles.")
        self.clear_fields()

    def calculate_totals(self, event):
        try:
            barril_50 = int(self.barril_50_entry.get()) if self.barril_50_entry.get() else 0
            barril_30 = int(self.barril_30_entry.get()) if self.barril_30_entry.get() else 0
            barril_20 = int(self.barril_20_entry.get()) if self.barril_20_entry.get() else 0
        except ValueError:
            self.total_label.config(text="Error")
            return

        total_barriles = barril_50 + barril_30 + barril_20
        self.total_label.config(text=str(total_barriles))

    def clear_fields(self):
        self.fecha_entry.delete(0, tk.END)
        self.fecha_entry.insert(0, datetime.datetime.now().strftime("%Y-%m-%d"))

        self.barril_50_entry.delete(0, tk.END)
        self.barril_30_entry.delete(0, tk.END)
        self.barril_20_entry.delete(0, tk.END)
        self.total_label.config(text="")

        self.enjuague_70_var.set(False)
        self.detergente_var.set(False)
        self.enjuague_frio_var.set(False)
        self.acido_fosforico_var.set(False)
        self.acido_peracetico_var.set(False)
        self.lanzas_var.set(False)

        self.responsable_entry.delete(0, tk.END)
        self.update_color()

    def sort_excel_data(self):
        data = list(self.ws.iter_rows(min_row=2, values_only=True))
        data.sort(key=lambda row: row[0].date() if isinstance(row[0], datetime.datetime) else row[0])

        for idx, row in enumerate(data, start=2):
            for col_idx, value in enumerate(row):
                self.ws.cell(row=idx, column=col_idx + 1, value=value)

if __name__ == "__main__":
    root = tk.Tk()
    app = BarrelCleaningApp(root)
    root.mainloop()
